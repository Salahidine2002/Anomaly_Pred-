import numpy as np
import openpyxl
import matplotlib.pyplot as plt
import os 
import pandas as pd
import scipy.io
from scipy.signal import detrend
from obspy.signal.detrend import polynomial
from pyts.image import MarkovTransitionField
from pyts.image import GramianAngularField
from time import time as tm
from sklearn.neighbors import NearestNeighbors 
from sklearn.preprocessing import MinMaxScaler
from copy import deepcopy

import torch
import torch.nn as nn
import torch.nn.functional as F
import torch.optim as optim
from torchvision import datasets, transforms
from torch.autograd import Variable
from torchvision.utils import save_image
from sklearn.model_selection import train_test_split
from torch.utils.data import Dataset, DataLoader, TensorDataset
from torchvision import transforms, datasets
from sklearn.metrics import accuracy_score, r2_score
from scipy import signal


def generate_wave(size):
    """
    Generates a wave function with random frequencies and amplitudes to simulate signal distortion.

    Args:
        size (int): Length of the signal to generate.

    Returns:
        numpy.ndarray: Array of wave values simulating signal distortion.
    """
    # Random frequencies and amplitudes for the wave
    freq1 = 0.0002 * np.random.rand()
    freq2 = 0.0001 * np.random.rand()
    freq3 = 0.000001 * np.random.rand()  # Unused frequency
    amp1 = np.random.rand()
    amp2 = np.random.rand()  # Unused amplitude

    distortion = np.zeros(size)
    for i in range(len(distortion)):
        # Generate wave values using sine and cosine functions
        distortion[i] = amp1 * np.sin(freq1 * i) * np.cos(freq2 * i)

    # Smooth the function using a median filter
    distortion = signal.medfilt(distortion, 3)
    # Normalize the values between 0 and 1, then scale
    distortion /= distortion.max()
    distortion *= 0.008
    return distortion


def load_data_brysthol_syn():
    """
    Generates synthetic Brysthol dataset for testing purposes.

    Returns:
        tuple: A tuple containing:
            - syn_data (numpy.ndarray): The synthetic profiles matrix of each cell in the stack.
            - cells (list): The names of the synthetic cells in the stack.
            - syn_time (numpy.ndarray): The synthetic signals sampling time.
    """
    syn_time = np.arange(0, 3400,  0.1/6)
    syn_data = []
    for i in range(75): 
        # Generate synthetic signals with random noise and voltage drops
        noise_mean = 1.4 if i == 1 else np.random.uniform(1.1, 1.4)
        noise_std = 0.01 if i == 1 else 0.005
        syn_signal = np.random.normal(noise_mean, noise_std, len(syn_time))

        # Add random voltage drops
        random_drops = np.random.uniform(0, len(syn_time), 10)
        drop_voltages = np.random.uniform(0, 0.8, len(random_drops))
        syn_signal[(random_drops).astype(int)] = drop_voltages

        # Add a linear trend and wave distortion
        syn_signal += np.linspace(0, 0.05, len(syn_signal)) + generate_wave(len(syn_signal))

        # Add voltage cutoffs
        cutoff_indices = [99999, 120000, 39999, 50000]
        cutoff_values = [0.15, 0.15, 0.6, 0.6]
        for index, value in zip(cutoff_indices, cutoff_values):
            syn_signal[index] = value
        syn_signal = np.concatenate((syn_signal[:100000], syn_signal[120000:]))
        syn_signal = np.concatenate((syn_signal[:40000], syn_signal[50000:]))

        syn_data.append(syn_signal) 

    syn_data = np.array(syn_data).T
    syn_time = np.concatenate((syn_time[:100000], syn_time[120000:]))
    syn_time = np.concatenate((syn_time[:40000], syn_time[50000:]))

    return syn_data, [f'Ch{str(i)}' for i in range(75)], syn_time

def random_sample(data, cells, failure, fail_index):
    """
    Randomly selects 5 cells and averages their profiles. If failure is indicated, includes a specific cell.

    Args:
        data (numpy.ndarray): The profiles matrix of each cell in the stack.
        cells (list): The names of the cells in the stack.
        failure (int): Indicates whether to simulate a failure (1) or normal condition (0).
        fail_index (int): The index of the failure signal in the data matrix.

    Returns:
        tuple: A tuple containing:
            - mock (numpy.ndarray): The average profile of the selected cells.
            - Indexes (list): The indices of the selected cells.
    """
    Indexes = [fail_index] if failure == 1 else []

    while len(Indexes) < 5:
        i = np.random.choice(list(range(len(cells))))
        if i not in Indexes:
            Indexes.append(i)
    mock = np.mean(data[:, Indexes], axis=1)

    return mock, Indexes

def generate_dataset(data, cells, size, fail_prop, fail_index):
    """
    Generates a dataset of mock profiles with a specified size and proportion of failures.

    Args:
        data (numpy.ndarray): The profiles matrix of each cell in the stack.
        cells (list): The names of the cells in the stack.
        size (int): The size of the generated dataset.
        fail_prop (float): The proportion of failures in the generated dataset.
        fail_index (int): The index of the failure signal in the data matrix.

    Returns:
        tuple: A tuple containing:
            - X (numpy.ndarray): The dataset of mock profiles.
            - y (numpy.ndarray): The labels vector indicating failures and normals.
            - composition (numpy.ndarray): The composition of cells used for each mock profile.
    """
    failure_size = int(fail_prop * size)
    success_size = size - failure_size

    X = np.zeros((size, data.shape[0]))
    y = np.array([1] * failure_size + [0] * success_size)
    composition = []
    for i in range(size):
        mock, idx = random_sample(data, cells, failure=y[i], fail_index=fail_index)
        X[i] = mock
        composition.append(idx)

    return X, y.astype(int), np.array(composition)

def chop_dataset(X, time, start, end):
    """
    Limits the time series to a certain duration before the failure to detect early signs.

    Args:
        X (numpy.ndarray): The dataset of mock profiles.
        time (numpy.ndarray): The time axis.
        start (float): The start time to chop the data.
        end (float): The end time to chop the data.

    Returns:
        tuple: A tuple containing:
            - X_chopped (numpy.ndarray): The dataset of clipped mock profiles.
            - time_chopped (numpy.ndarray): The time axis for the clipped profiles.
    """
    mask = (time >= start) & (time <= end)
    X_chopped = X[:, mask]
    time_chopped = time[mask]
    return X_chopped, time_chopped

def normalize(X):
    """
    Normalizes the dataset to eliminate amplitude differences between signals.

    Args:
        X (numpy.ndarray): The dataset of chopped mock profiles.

    Returns:
        numpy.ndarray: The normalized dataset.
    """
    return np.array([X[i] / X[i][100] for i in range(len(X))])

def filter(X, time, plot=True):
    """
    Eliminates outliers from the dataset based on a Z-score threshold.

    Args:
        X (numpy.ndarray): The input dataset.
        time (numpy.ndarray): The time axis for the signals.
        plot (bool, optional): Whether to plot the Z-score over time. Defaults to True.

    Returns:
        tuple: A tuple containing:
            - X_filt (numpy.ndarray): The filtered dataset.
            - time_filt (numpy.ndarray): The time axis for the filtered signals.
    """
    mean = np.mean(X[1, :])
    std = np.std(X[1, :])
    Z_score = (X[1, :] - mean) / std
    threshold = 3
    filt = np.abs(Z_score) < threshold

    X_filt = X[:, filt]
    time_filt = time[filt]

    if plot:
        plt.plot(time, Z_score)
        plt.xlabel('time')
        plt.ylabel('Sigma score')
        plt.hlines(-threshold, min(time_filt), max(time_filt), colors='r')
        plt.hlines(threshold, min(time_filt), max(time_filt), colors='r')
        plt.show()

    return X_filt, time_filt

def fill_signal(X, time, time_diff_threshold=1, size=1000):
    """
    Fills missing regions in the signal based on a time difference threshold.

    Args:
        X (numpy.ndarray): The input dataset.
        time (numpy.ndarray): The time axis for the signals.
        time_diff_threshold (float, optional): The threshold for detecting missing regions. Defaults to 1.
        size (int, optional): The size of the window for noise calculation. Defaults to 1000.

    Returns:
        tuple: A tuple containing:
            - X_new (numpy.ndarray): The dataset with filled missing regions.
            - time_new (list): The time axis for the new filled signals.
    """
    diff = np.abs(time[1:] - time[:-1])
    indexes = np.where((diff > time_diff_threshold))[0]
    if len(indexes) == 0:
        return X, time

    new_lengths = []
    time_new = []
    i = 0
    k = 0
    while i < indexes[-1]:
        j = indexes[k]
        time_new += list(time[i:j])
        i = j
        j = indexes[k] + 1
        length = int((5 / 4) * (abs((time[j] - time[i]) / (np.mean(diff)))))
        new_lengths.append(length)
        time_new += list(np.linspace(time[i], time[j], length))
        i = j
        k += 1

    time_new += list(time[i:])
    X_new = []
    for l in range(len(X)):
        i = 0
        k = 0
        Signal_new = []
        while i < indexes[-1]:
            j = indexes[k]
            Signal_new += list(X[l, i:j])
            i = j
            j = indexes[k] + 1
            noise_before_std = np.std(X[l, i - size:i])
            noise_after_std = np.std(X[l, j:j + size])
            noise_before_mean = np.mean(X[l, i - size:i])
            noise_after_mean = np.mean(X[l, j:j + size])
            Signal_new += list(np.random.normal((noise_after_mean + noise_before_mean) / 2, (noise_before_std + noise_after_std) / 2, new_lengths[k]))
            i = j
            k += 1

        Signal_new += list(X[l, i:])
        X_new.append(Signal_new)

    return np.array(X_new), time_new

def linear_detrend(X):
    """
    Eliminates the first-order trend from input signals.

    Args:
        X (numpy.ndarray): The input dataset.

    Returns:
        numpy.ndarray: The detrended dataset.
    """
    X_dtr = np.zeros(X.shape)
    for i in range(len(X)):
        signal = np.copy(X[i, :])
        X_dtr[i, :] = polynomial(signal, order=1)

    return X_dtr

def plot_s(time_chopped, X_chopped, X_norm, time_filt, X_filt, time_new, X_new, X_dtr):
    """
    Plots various stages of signal preprocessing for comparison.

    Args:
        time_chopped (numpy.ndarray): The time axis for the chopped signals.
        X_chopped (numpy.ndarray): The dataset of chopped signals.
        X_norm (numpy.ndarray): The dataset of normalized signals.
        time_filt (numpy.ndarray): The time axis for the filtered signals.
        X_filt (numpy.ndarray): The dataset of filtered signals.
        time_new (list): The time axis for the filled signals.
        X_new (numpy.ndarray): The dataset of filled signals.
        X_dtr (numpy.ndarray): The dataset of detrended signals.
    """
    indx = 10
    fig, axs = plt.subplots(1, 5, figsize=(15, 3))

    axs[0].set_ylim((0.8, 1.4))
    axs[0].plot(time_chopped, X_chopped[indx, :])
    axs[0].title.set_text('Chopped signal')
    axs[0].set_ylabel('voltage (v)')
    axs[0].set_xlabel('time (h)')

    axs[1].set_ylim((0.6, 1.1))
    axs[1].plot(time_chopped, X_norm[indx, :])
    axs[1].title.set_text('Normalized')
    axs[1].set_xlabel('time (h)')

    axs[2].set_ylim((0.6, 1.1))
    axs[2].plot(time_filt, X_filt[indx, :])
    axs[2].title.set_text('Filtered')
    axs[2].set_xlabel('time (h)')

    axs[3].set_ylim((0.6, 1.1))
    axs[3].plot(time_new, X_new[indx, :])
    axs[3].title.set_text('Filled')
    axs[3].set_xlabel('time (h)')

    axs[4].set_ylim((-0.4, 0.1))
    axs[4].plot(time_new, X_dtr[indx, :])
    axs[4].title.set_text('Detrended')
    axs[4].set_xlabel('time (h)')
    plt.show()

def noise_evo(X, window_size, step, time=[None]):
    """
    Computes the noise standard deviation evolution for every signal in the dataset using a sliding window.

    Args:
        X (numpy.ndarray): The input dataset.
        window_size (int): The sliding window size to compute the noise standard deviation.
        step (int): The stride of the sliding window.
        time (list, optional): The time axis. Defaults to [None].

    Returns:
        numpy.ndarray: The noise evolution signals. If time is provided, also returns the time frames.
    """
    i = 0
    noise_list = []
    time_frames = []

    while True:
        noise = np.std(X[:, i:i + window_size], axis=1)
        noise_list.append(noise)
        if time[0] is not None:
            time_frames.append(np.mean(time[i:i + window_size]))
        i += step
        if i + window_size > X.shape[1]:
            break

    noise_list.append(np.std(X[:, i:], axis=1))
    if time[0] is None:
        return np.array(noise_list).T
    else:
        time_frames.append(np.mean(time[i:]))
        return np.array(noise_list).T, time_frames
    
def preprocess_random_brysthol(Indexes, params):
    """
    Preprocesses the Brysthol dataset by simulating the Grenoble dataset's sampling rate and generating sub-signals
    to prevent model overfitting. It involves chopping, downsampling, filtering, generating mock profiles, filling gaps,
    detrending, and computing the noise evolution of each signal.

    Args:
        Indexes (list): Indices of signals to consider.
        params (dict): Dictionary containing parameters for preprocessing, such as 'low', 'high', 'size', 'prop', 'fail_id', and 'window_size'.

    Returns:
        tuple: A tuple containing:
            - X_noise (numpy.ndarray): The noise evolution of each preprocessed signal.
            - y_final (numpy.ndarray): The labels of each signal.
            - sample_id (numpy.ndarray): The origin index of each signal cut.
    """

    # Load Brysthol Dataset
    data, cells, time = load_data_brysthol_syn()

    # Chop the signals to the desired length
    X_chopped, time_chopped = chop_dataset(np.array(data).T, time, params['low'], params['high'])

    # Downsample the signals to mimic Grenoble sampling rate
    NN = NearestNeighbors(n_neighbors=10)
    NN.fit(time_chopped.reshape(-1, 1))
    sub_samples = np.sort(np.unique(time_chopped.astype(int)))        # Take only one sample every hour 
    sub_samples_neighb = NN.kneighbors(sub_samples.reshape(-1, 1))[1] # Find the 10 nearest neighbors to each hourly sample 
    X_sampled = np.array([np.mean(X_chopped[i][sub_samples_neighb], axis=1) for i in range(len(X_chopped))]) # average

    # Filtering the outliers before mock profile generation to optimize the calculations
    X_filt, time_filt = filter(X_sampled, sub_samples, plot=False)

    # Generating mock profiles dataset 
    X_gen, y, composition = generate_dataset(X_filt.T, Indexes, params['size'], params['prop'], params['fail_id'])

    # Filling the missing gaps of the signals 
    X_new, time_new = fill_signal(X_gen, time_filt, size=20, time_diff_threshold=2.5)

    # Detrending the signals 
    X_dtr = linear_detrend(X_new)

    # Sampling 5 smaller cuts of each signal with random start points (to prevent overfitting on one interval)
    max_start = X_dtr.shape[1]-params['window_size']
    rand_start = np.random.randint(0, max_start, (X_dtr.shape[0], 5))
    X_final = []
    y_final = []
    sample_id = []
    for i in range(X_dtr.shape[0]) : 
        for j in range(5) : 
            X_final.append(X_dtr[i, rand_start[i, j]:rand_start[i, j]+params['window_size']])
            y_final.append(y[i])
            sample_id.append(composition[i])  # Documenting where the signal came from 
    X_final  = np.array(X_final)

    # Computing the noise evolution of each signal
    X_noise = noise_evo(X_final, window_size=30, step=8) 

    return X_noise, np.array(y_final), np.array(sample_id)    ## TODO : exploit the informarion of sample_id ? 

def transform_mtx(X_noise, steps, params):
    """
    Transforms the signal's noise evolution into images using GASF, GADF, and MTF transformations.

    Args:
        X_noise (numpy.ndarray): The noise evolution of each preprocessed signal.
        steps (list): The points in the signals to consider for the image transformations.
        params (dict): Dictionary containing transformation parameters, such as 'im_size'.

    Returns:
        numpy.ndarray: An array of shape (nbr_signals, 3, image_size, image_size) containing the transformed images.
    """

    X_in = X_noise[:, steps]

    # Compute Gramian summation angular fields
    gasf = GramianAngularField(method='summation', image_size=params['im_size'])
    X_gasf = gasf.fit_transform(X_in)

    # Compute Gramian difference angular fields
    gadf = GramianAngularField(method='difference', image_size=params['im_size'])
    X_gadf = gadf.fit_transform(X_in)

    # Compute Markov Transition Field
    mtf = MarkovTransitionField(image_size=params['im_size'], n_bins=4)
    X_mtf = mtf.fit_transform(X_in)

    X_vae = np.zeros((len(X_gadf), 3, params['im_size'], params['im_size']))

    X_vae[:, 0, :, :] = X_gadf
    X_vae[:, 1, :, :] = X_gasf
    X_vae[:, 2, :, :] = X_mtf

    return X_vae

def load_data_grenoble_syn():
    """
    Generates synthetic Grenoble dataset for testing purposes.

    Returns:
        tuple: A tuple containing:
            - syn_data (numpy.ndarray): The synthetic profiles matrix of each cell in the stack.
            - syn_time (numpy.ndarray): The synthetic signals sampling time.
    """
    syn_time = np.arange(0, 856,  856/755)
    syn_data = []
    for i in range(75): 
        # Generating random signals 
        if i==1 : 
            noise_mean = 1.4
            noise_std = 0.01
        else : 
            noise_mean = np.random.uniform(1, 1.3)
            noise_std = 0.0045
        syn_signal = np.random.normal(noise_mean, noise_std, len(syn_time))

        # Adding random voltage drops
        random_drops = np.random.uniform(0, len(syn_time), 10)
        drop_voltages = np.random.uniform(0, 0.8, len(random_drops))
        syn_signal[(random_drops).astype(int)] = drop_voltages

        # Adding a linear trend to the signals 
        syn_signal = syn_signal + np.linspace(0, 0.05, len(syn_signal)) + generate_wave(len(syn_signal))

        # Adding voltage cutoffs (same locations for all signals if the stack) 
        syn_signal[99] = 0.15
        syn_signal[120] = 0.15
        syn_signal = np.concatenate((syn_signal[:100], syn_signal[120:]))
        syn_signal[399] = 0.6
        syn_signal[450] = 0.6
        syn_signal = np.concatenate((syn_signal[:400], syn_signal[450:]))

        syn_data.append(syn_signal) 

    syn_data = np.array(syn_data)
    syn_data = syn_data.T

    syn_time = np.concatenate((syn_time[:100], syn_time[120:]))
    syn_time = np.concatenate((syn_time[:400], syn_time[450:]))

    return syn_data, syn_time

def load_data_grenoble():
    """
    Loads the Grenoble dataset from a specified path.

    Returns:
        tuple: A tuple containing:
            - X (numpy.ndarray): The profiles matrix of each cell in the stack.
            - time (numpy.ndarray): The signals sampling time.
    """
    path_data = '/dbfs/FileStore/tables/GrenobleData/20230417_MO_GVFR_000001211_BOU_EHT_Durability_Results_10_3Nml.xlsx'
    workbook = openpyxl.load_workbook(path_data)
    sheet_names = workbook.sheetnames
    stack_data_name = 'Complete Data'
    stack_data = workbook[stack_data_name]
    workbook.close()

    data = []
    for row in stack_data.iter_rows(values_only=True):
        data.append(row)

    df = pd.DataFrame(data)
    rename = {} 
    for i in range(len(df.columns)) : 
        rename[i] = df.loc[0, i]

    df.rename(columns=rename, inplace=True)
    df = df.drop(0)

    df['Time Stamp'] = pd.to_datetime(df['Time Stamp'])
    df['time_seconds'] = (df['Time Stamp'] - pd.Timestamp("2023-04-17 08:14:04")) // pd.Timedelta(seconds=1)

    Columns = {}
    for s in range(1, 4) : 
        for n in range(1, 26) : 
            Columns[f'SS{s}_Cell{n}_Voltage'] = s*n 

    X, time = df[list(Columns.keys())], np.array(df['time_seconds'])
    stop_idx = 800
    X, time = np.array(X[45:800]), time[45:800]

    time /= 3600
    time -= time[0]

    return X, time

def preprocess_random_grenoble(params):
    """
    Preprocesses the Grenoble dataset by filtering, generating mock profiles, filling gaps, detrending, and extracting sub-signals.
    This function aims to prepare the data for further analysis or model training.

    Args:
        params (dict): Dictionary containing parameters for preprocessing, such as 'size', 'prop', 'fail_id', and 'window_size'.

    Returns:
        tuple: A tuple containing:
            - X_noise (numpy.ndarray): The noise evolution of each preprocessed signal.
            - sample_id (numpy.ndarray): The origin index of each signal cut.
            - composition (numpy.ndarray): The composition of cells used for each mock profile.
    """
    # Loading Grenoble data 
    X, time = load_data_grenoble_syn()
    data, cells = X, list(range(0, 75))

    # Filtering data
    X_filt, time_filt = filter(data.T, time, plot=False)

    # Generating mock profiles
    X_gen, _, composition = generate_dataset(X_filt.T, cells, params['size'], params['prop'], params['fail_id'])

    # Filling missing gaps 
    X_new, time_new = fill_signal(X_gen, time_filt, time_diff_threshold=10, size=30)

    # Detrending signals 
    X_dtr = linear_detrend(X_new)

    # Extracting sub-signals
    max_start = X_dtr.shape[1]-params['window_size']
    rand_start = np.random.randint(0, max_start, (X_dtr.shape[0], 5))
    X_final = []
    starts = []
    sample_id = []
    for i in range(X_dtr.shape[0]) : 
        for j in range(5) : 
            X_final.append(X_dtr[i, rand_start[i, j]:rand_start[i, j]+params['window_size']])
            starts.append(rand_start[i, j])
            sample_id.append(composition[i])
    X_final  = np.array(X_final)

    # Computing Noise evolution signals
    X_noise = noise_evo(X_final, window_size=30, step=8) 

    return X_noise, np.array(sample_id), np.array(composition)

def process_combined_data(X_noise_source, X_noise_target, params):
    """
    Processes combined data from two sources by concatenating, scaling, and transforming the noise evolution signals into images.

    Args:
        X_noise_source (numpy.ndarray): The noise evolution signals from the source dataset.
        X_noise_target (numpy.ndarray): The noise evolution signals from the target dataset.
        params (dict): Dictionary containing parameters for the transformation, such as 'im_size'.

    Returns:
        tuple: A tuple containing:
            - X_noise_source_scaled (numpy.ndarray): Scaled noise evolution signals from the source dataset.
            - X_noise_target_scaled (numpy.ndarray): Scaled noise evolution signals from the target dataset.
            - X_vae_source (numpy.ndarray): Transformed images from the source dataset.
            - X_vae_target (numpy.ndarray): Transformed images from the target dataset.
    """
    n = len(X_noise_source)
    X_noise = np.concatenate((X_noise_source, X_noise_target), axis=0)

    scaler = MinMaxScaler(feature_range=(0, 1))
    X_noise = np.float32(scaler.fit_transform(X_noise))

    steps = list(range(X_noise.shape[1]))
    X_vae = np.float32(transform_mtx(X_noise, steps, params))
    X_vae[:, :2, :, :] = np.float32((X_vae[:, :2, :, :]+1)/2)
    
    return X_noise[:n], X_noise[n:], X_vae[:n], X_vae[n:]

def array_to_loader(X, y, dim=61):
    """
    Converts arrays into a DataLoader for PyTorch models.

    Args:
        X (numpy.ndarray): Input features.
        y (numpy.ndarray): Target labels.
        dim (int, optional): Dimension of the target labels. Defaults to 61.

    Returns:
        DataLoader: A DataLoader object containing the dataset.
    """
    X_to = torch.from_numpy(X.astype(np.float32))
    y_to = torch.from_numpy(y.astype(np.float32))
    y_to = y_to.view(-1, dim)
    dataset = TensorDataset(X_to, y_to)
    loader = DataLoader(dataset, batch_size=64, shuffle=True, num_workers=2, pin_memory=True)
    return loader

def investigate(vae, loader):
    """
    Investigates the embeddings and reconstructions produced by a VAE model.

    Args:
        vae (Model): The VAE model to investigate.
        loader (DataLoader): DataLoader containing the dataset to investigate.

    Returns:
        tuple: A tuple containing:
            - embedding (numpy.ndarray): Embeddings produced by the VAE.
            - y_emb_ (numpy.ndarray): Original labels of the embeddings.
            - y_pred_ (numpy.ndarray): Predicted labels by the VAE.
            - reconst (numpy.ndarray): Reconstructions produced by the VAE.
            - embedding_samples (numpy.ndarray): Sampled embeddings from the latent space.
    """
    embedding = []
    y_emb_ = []
    y_pred_ = []
    reconst = []
    embedding_samples = []
    for data, labels in loader:
        X_reconst, mu, _, out, z = vae(data)
        embedding += list(mu.detach().numpy())
        embedding_samples += list(z.detach().numpy())
        y_emb_ += list(labels.detach().numpy()[:, 0])
        y_pred_ += list(out.detach().numpy())
        reconst += list(np.concatenate((labels.detach().numpy()[:, 1:Signal_dim+1], X_reconst.detach().numpy()), axis=1))
    embedding_samples = np.array(embedding_samples)
    embedding = np.array(embedding)
    y_emb_ = np.array(y_emb_)
    y_pred_ = np.array(y_pred_)
    reconst = np.array(reconst)
    
    return embedding, y_emb_, y_pred_, reconst, embedding_samples